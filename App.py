import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
import pathlib
from openpyxl import Workbook

#Visual do app

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.config_layout()
        self.sistema()

    def config_layout(self):
        self.title("Cadastro")
        self.geometry("720x360")

    def sistema(self):
        barra = ctk.CTkFrame(self, width=720, height=5, corner_radius=0, bg_color="purple2", fg_color="purple2").place(x=0, y=50)
        titulo = ctk.CTkLabel(self, text="Cadastro Clientes", text_color=["#000", "#fff"], font=("Impact",24)).place(x=30, y=20)

        dataxls = pathlib.Path('cadastros.xlsx')

        if dataxls.exists():
            pass
        else:
            dataxls=Workbook()
            abaxls = dataxls.active
            abaxls['A1']="NOME"
            abaxls['B1']="CPF"
            abaxls['C1']="CONTATO"
            abaxls['D1']="CARRO"
            abaxls['E1']="COMBUSTIVEL"

            dataxls.save("cadastros.xlsx")

        def enviar_dados():

            nome = valor_nome.get()
            cpf = valor_cpf.get()
            celular = valor_celular.get()
            carro = valor_carro.get()
            combustivel = valor_combustivel.get()

            dataxls = openpyxl.load_workbook('cadastros.xlsx')
            abaxls = dataxls.active
            abaxls.cell(column=1, row=abaxls.max_row+1, value=nome)
            abaxls.cell(column=2, row=abaxls.max_row, value=cpf)
            abaxls.cell(column=3, row=abaxls.max_row, value=celular)
            abaxls.cell(column=4, row=abaxls.max_row, value=carro)
            abaxls.cell(column=5, row=abaxls.max_row, value=combustivel)

            dataxls.save(r"cadastros.xlsx")

            messagebox.showinfo("Sistema", "Obrigado pelo seu cadastro")

        def clear():
            valor_nome.set("")
            valor_cpf.set("")
            valor_celular.set("")
            valor_carro.set("")
            valor_combustivel.set("")


        #Variaveis
        valor_nome = StringVar()
        valor_cpf = StringVar()
        valor_celular = StringVar()
        valor_carro = StringVar()
        valor_combustivel = StringVar()

        #entries
        nome_entry = ctk.CTkEntry(self, width = 350, textvariable=valor_nome,font = ("Century Gothic", 16), fg_color="transparent")
        cpf_entry = ctk.CTkEntry(self,width = 200, textvariable=valor_cpf, font = ("Century Gothic", 16), fg_color="transparent")
        celular_entry = ctk.CTkEntry(self,width = 200, textvariable=valor_celular, font = ("Century Gothic", 16), fg_color="transparent")
        carro_entry = ctk.CTkEntry(self,width = 350, textvariable=valor_carro, font = ("Century Gothic", 16), fg_color="transparent")
        combustivel_entry = ctk.CTkEntry(self,width = 350, textvariable=valor_combustivel, font = ("Century Gothic", 16), fg_color="transparent")

        #campos
        var_nome = ctk.CTkLabel(self, text="Nome",text_color=["#000", "#fff"], font=("Century Gothic Bold",16))
        var_cpf = ctk.CTkLabel(self, text="CPF", text_color=["#000", "#fff"], font=("Century Gothic Bold",16))
        var_celular = ctk.CTkLabel(self, text="Contato", text_color=["#000", "#fff"], font=("Century Gothic Bold",16))
        var_carro = ctk.CTkLabel(self, text="Carro", text_color=["#000", "#fff"], font=("Century Gothic Bold",16))
        var_combustivel = ctk.CTkLabel(self, text="Combustível", text_color=["#000", "#fff"], font=("Century Gothic Bold",16))

        botao_enviar = ctk.CTkButton(self, text="Cadastrar".upper(), command=enviar_dados, fg_color="#339", hover_color="#335")
        botao_limpar = ctk.CTkButton(self, text="Limpar".upper(), command=clear, fg_color="#111", hover_color="#111")

        #posições
        var_nome.place(x=50, y=70)
        nome_entry.place(x=50, y=100)

        var_carro.place(x=50, y=150)
        carro_entry.place(x=50, y=180)

        var_cpf.place(x=450, y=70)
        cpf_entry.place(x=450, y=100)

        var_celular.place(x=450, y=150)
        celular_entry.place(x=450, y=180)

        var_combustivel.place(x=50, y=230)
        combustivel_entry.place(x=50, y=260)

        botao_enviar.place(x=500, y=280)
        botao_limpar.place(x=500, y=240)
    

if __name__=="__main__":
    app = App()
    app.mainloop()